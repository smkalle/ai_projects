"""Export functionality for extracted data."""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
import json
import csv
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config.settings import settings


logger = logging.getLogger(__name__)


class ExportManager:
    """Manage data export in various formats."""
    
    def __init__(self):
        self.export_dir = settings.export_dir
        self._ensure_directories()
    
    def _ensure_directories(self):
        """Ensure export directory exists."""
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def _generate_filename(self, format: str) -> Path:
        """Generate unique filename for export."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"langextract_export_{timestamp}.{format}"
        return self.export_dir / filename
    
    def export_to_csv(self, 
                      data: List[Dict[str, Any]], 
                      include_metadata: bool = True) -> Path:
        """Export data to CSV format."""
        try:
            # Flatten data for CSV export
            flattened_data = []
            
            for item in data:
                if "extractions" in item:
                    for extraction in item["extractions"]:
                        row = {
                            "source_file": item.get("source_file", ""),
                            "template_used": item.get("template_used", ""),
                            "extraction_id": item.get("id", "")
                        }
                        
                        if include_metadata:
                            row.update({
                                "processing_time": item.get("processing_time", ""),
                                "timestamp": item.get("timestamp", ""),
                                "token_count": item.get("token_count", "")
                            })
                        
                        # Flatten extraction data
                        if isinstance(extraction, dict):
                            for key, value in extraction.items():
                                if isinstance(value, (list, dict)):
                                    row[key] = json.dumps(value)
                                else:
                                    row[key] = value
                        
                        flattened_data.append(row)
            
            # Write to CSV
            file_path = self._generate_filename("csv")
            
            if flattened_data:
                df = pd.DataFrame(flattened_data)
                df.to_csv(file_path, index=False, encoding='utf-8-sig')
                logger.info(f"CSV export completed: {file_path}")
            else:
                # Write empty CSV with headers
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(["source_file", "template_used", "extraction_id"])
            
            return file_path
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            raise
    
    def export_to_json(self, 
                       data: List[Dict[str, Any]], 
                       pretty: bool = True) -> Path:
        """Export data to JSON format."""
        try:
            file_path = self._generate_filename("json")
            
            with open(file_path, 'w', encoding='utf-8') as f:
                if pretty:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                else:
                    json.dump(data, f, ensure_ascii=False)
            
            logger.info(f"JSON export completed: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            raise
    
    def export_to_excel(self, 
                        data: List[Dict[str, Any]], 
                        include_metadata: bool = True) -> Path:
        """Export data to Excel format with formatting."""
        try:
            file_path = self._generate_filename("xlsx")
            
            # Create workbook
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = []
                for item in data:
                    summary_data.append({
                        "Source File": item.get("source_file", ""),
                        "Template": item.get("template_used", ""),
                        "Extractions": item.get("extraction_count", len(item.get("extractions", []))),
                        "Processing Time (s)": round(item.get("processing_time", 0), 2),
                        "Timestamp": item.get("timestamp", "")
                    })
                
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # Extractions sheet
                all_extractions = []
                for item in data:
                    source = item.get("source_file", "")
                    template = item.get("template_used", "")
                    
                    for idx, extraction in enumerate(item.get("extractions", [])):
                        extraction_row = {
                            "Source": source,
                            "Template": template,
                            "Index": idx + 1
                        }
                        
                        # Flatten extraction data
                        if isinstance(extraction, dict):
                            for key, value in extraction.items():
                                if isinstance(value, (list, dict)):
                                    extraction_row[key] = json.dumps(value, ensure_ascii=False)
                                else:
                                    extraction_row[key] = value
                        
                        all_extractions.append(extraction_row)
                
                if all_extractions:
                    extractions_df = pd.DataFrame(all_extractions)
                    extractions_df.to_excel(writer, sheet_name="Extractions", index=False)
                
                # Raw data sheet (if metadata included)
                if include_metadata:
                    raw_df = pd.DataFrame(data)
                    raw_df.to_excel(writer, sheet_name="Raw Data", index=False)
            
            # Apply formatting
            self._format_excel(file_path)
            
            logger.info(f"Excel export completed: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise
    
    def _format_excel(self, file_path: Path):
        """Apply formatting to Excel file."""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze header row
                ws.freeze_panes = "A2"
            
            wb.save(file_path)
            
        except Exception as e:
            logger.warning(f"Excel formatting failed: {e}")
    
    def export_to_latex(self, 
                        data: List[Dict[str, Any]], 
                        table_caption: str = "Extracted Medical Data") -> Path:
        """Export data to LaTeX table format."""
        try:
            file_path = self._generate_filename("tex")
            
            latex_content = [
                "\\documentclass{article}",
                "\\usepackage{booktabs}",
                "\\usepackage{longtable}",
                "\\usepackage[utf8]{inputenc}",
                "\\begin{document}",
                "",
                f"\\begin{{longtable}}{{|l|l|l|p{{6cm}}|}}",
                f"\\caption{{{table_caption}}} \\\\",
                "\\hline",
                "\\textbf{Source} & \\textbf{Template} & \\textbf{Field} & \\textbf{Value} \\\\",
                "\\hline",
                "\\endfirsthead",
                "\\multicolumn{4}{c}{{\\tablename\\ \\thetable{} -- continued from previous page}} \\\\",
                "\\hline",
                "\\textbf{Source} & \\textbf{Template} & \\textbf{Field} & \\textbf{Value} \\\\",
                "\\hline",
                "\\endhead",
                "\\hline \\multicolumn{4}{r}{{Continued on next page}} \\\\",
                "\\endfoot",
                "\\hline",
                "\\endlastfoot"
            ]
            
            # Add data rows
            for item in data:
                source = self._escape_latex(item.get("source_file", ""))
                template = self._escape_latex(item.get("template_used", ""))
                
                for extraction in item.get("extractions", []):
                    if isinstance(extraction, dict):
                        for field, value in extraction.items():
                            field_escaped = self._escape_latex(str(field))
                            value_escaped = self._escape_latex(str(value)[:100])  # Limit length
                            latex_content.append(
                                f"{source} & {template} & {field_escaped} & {value_escaped} \\\\"
                            )
                            latex_content.append("\\hline")
            
            latex_content.extend([
                "\\end{longtable}",
                "",
                "\\end{document}"
            ])
            
            # Write to file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(latex_content))
            
            logger.info(f"LaTeX export completed: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"LaTeX export failed: {e}")
            raise
    
    def _escape_latex(self, text: str) -> str:
        """Escape special LaTeX characters."""
        replacements = {
            '\\': '\\textbackslash{}',
            '{': '\\{',
            '}': '\\}',
            '$': '\\$',
            '&': '\\&',
            '#': '\\#',
            '^': '\\textasciicircum{}',
            '_': '\\_',
            '~': '\\textasciitilde{}',
            '%': '\\%'
        }
        
        for char, replacement in replacements.items():
            text = text.replace(char, replacement)
        
        return text
    
    def export_to_bibtex(self, 
                         data: List[Dict[str, Any]]) -> Path:
        """Export literature review data to BibTeX format."""
        try:
            file_path = self._generate_filename("bib")
            
            bibtex_entries = []
            entry_count = 0
            
            for item in data:
                if item.get("template_used") == "literature_review":
                    for extraction in item.get("extractions", []):
                        if isinstance(extraction, dict):
                            entry_count += 1
                            
                            # Generate entry key
                            authors = extraction.get("authors", "Unknown")
                            year = extraction.get("year", "0000")
                            entry_key = f"{authors.split()[0]}{year}"
                            
                            # Build BibTeX entry
                            entry = [f"@article{{{entry_key},"]
                            
                            if "title" in extraction:
                                entry.append(f'  title = {{{extraction["title"]}}},')
                            if "authors" in extraction:
                                entry.append(f'  author = {{{extraction["authors"]}}},')
                            if "year" in extraction:
                                entry.append(f'  year = {{{extraction["year"]}}},')
                            if "journal" in extraction:
                                entry.append(f'  journal = {{{extraction["journal"]}}},')
                            if "findings" in extraction:
                                entry.append(f'  abstract = {{{extraction["findings"]}}},')
                            
                            entry.append("}")
                            bibtex_entries.append('\n'.join(entry))
            
            # Write to file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\n\n'.join(bibtex_entries))
            
            logger.info(f"BibTeX export completed: {file_path} ({entry_count} entries)")
            return file_path
            
        except Exception as e:
            logger.error(f"BibTeX export failed: {e}")
            raise
    
    def get_mime_type(self, format: str) -> str:
        """Get MIME type for download."""
        mime_types = {
            "csv": "text/csv",
            "json": "application/json",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "tex": "text/x-tex",
            "latex": "text/x-tex",
            "bib": "text/x-bibtex",
            "bibtex": "text/x-bibtex"
        }
        
        return mime_types.get(format.lower(), "application/octet-stream")